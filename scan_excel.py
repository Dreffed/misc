from utils import load_pickle, save_pickle
import os
import openpyxl
import logging
from logging.config import fileConfig

def process_sheet(ws):
    data = {}
    for rowNum in range(1, ws.max_row):
        for colNum in range(1, ws.max_column):
            cell = ws.cell(row=rowNum, column=colNum)
            if cell.value:
                if isinstance(cell.value, str):
                    if cell.value[0] == "=":
                        cell_ref = "{}[{},{}]".format(ws.title, colNum, rowNum)
                        data[cell_ref] = cell.value

    logger.info("\t{} => {} formula cells".format(ws.title, len(data)))
    return data
                
def process_workbook(filename):
    data = {}

    wb = openpyxl.load_workbook(filename)
    logger.info("Opened {}".format(type(wb)))

    for name in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(name)
        data = {**data, **process_sheet(ws)}

    return data

fileConfig('logging_config.ini')
logger = logging.getLogger(__name__)
homepath = os.path.expanduser("~")
filepaths = ['OneDrive - Great Canadian Railtour Co','AllSell']
filename = "{}.xlsx".format('2020_ALL_SELL_v1.3_MAIN TOOL - UPDATE FIRST')
folderpath = os.path.join(homepath, *filepaths)
filepath = os.path.join(folderpath, filename)
data = process_workbook(filepath)
picklename = 'Analysis_{}.pickle'.format(filename)
save_pickle(data=data, picklename=picklename)

all_sell__config = {
    "data":{
        "CDN":{
            "fieldmap":{
                [
                    {
                        "name": "",
                        "column": "",
                        "field": "",
                        "value": 0,
                        "formula": "",
                        "condition":"",
                        "format": "%d/%m/%Y",
                        "type": "date"
                        "border": "default",
                        "fill": "default",
                        "font": "default"
                    },

                ]
            },
            "offsets":{
                "row": 0,
                "column": 0
            }
        }
    },
    "styles":{
        "fills":{
            "default": {
                "fill_type": "solid",
                "start_color": "FFFFCC"
            }
        },
        "border": {
            "default": {
                "color": "B2B2B2",
                "style": "thin"
            }
        },
        "fonts": {
            "default": {
                "name": "Ariel",
                "size": 10
            }
        }              
    }
}



